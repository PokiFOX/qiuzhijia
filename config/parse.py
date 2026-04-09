import json
import openpyxl
import mysql.connector
import requests

from tapah import const
from tapah import data
from tapah import function
from tapah import reserved

data.mysql_conn = mysql.connector.connect(
	host		= reserved.mysql_host,
	port		= reserved.mysql_port,
	user		= reserved.mysql_username,
	password	= reserved.mysql_password,
	database	= reserved.mysql_database,
	auth_plugin = 'caching_sha2_password',
)
data.mysql_conn.connect()
zone = []
sector = []
level = []
field = []

def parse_global(page):
	global zone, sector, level
	for i in range(2, page.max_row + 1):
		cell_value = function.getcell_str(page, i, 1)
		if cell_value == "": continue
		zone.append(cell_value)

	for i in range(2, page.max_row + 1):
		cell_value = function.getcell_str(page, i, 3)
		if cell_value == "": continue
		sector.append(cell_value)

	for i in range(2, page.max_row + 1):
		cell_value = function.getcell_str(page, i, 2)
		if cell_value == "": continue
		level.append(cell_value)

def parse_field(page):
	global field
	for i in range(2, page.max_row + 1):
		name = function.getcell_str(page, i, 1)
		field_item = next((item for item in field if item['name'] == name), None)
		mapname = function.getcell_str(page, i, 2)
		if field_item is None:
			type = function.getcell_str(page, i, 3)
			star = function.getcell_int(page, i, 4)
			content = function.getcell_str(page, i, 5)
			field.append({
				"name": name,
				"type": type,
				"star": star,
				"content": content,
				"mapname": [mapname],
			})
		else:
			field_item['mapname'].append(mapname)

def upload_config():
	try:
		print(const.url_setzone)
		r = requests.post(function.url(const.url_setzone), json = zone, headers = const.request_headers, timeout = 15)
		r.raise_for_status()
		print(const.url_setsector)
		r = requests.post(function.url(const.url_setsector), json = sector, headers = const.request_headers, timeout = 15)
		r.raise_for_status()
		print(const.url_setlevel)
		r = requests.post(function.url(const.url_setlevel), json = level, headers = const.request_headers, timeout = 15)
		r.raise_for_status()
		print(const.url_setfield)
		r = requests.post(function.url(const.url_setfield), json = field, headers = const.request_headers, timeout = 15)
		r.raise_for_status()
	except requests.RequestException as err:
		print(f"设置失败 : {err}")

def parse_enterprise(page):
	for row in range(2, page.max_row + 1):
		enterprise = {
			"zone": function.getcell_str(page, row, const.column_zone),				# 地区
			"city": function.getcell_str(page, row, const.column_city),				# 城市
			"name": function.getcell_str(page, row, const.column_name),				# 公司名称
			"shortname": function.getcell_str(page, row, const.column_short),		# 公司简称
			"brief": function.getcell_str(page, row, const.column_brief),			# 公司简介
			"upper": function.getcell_str(page, row, const.column_upper),			# 上级单位
			"sector": function.getcell_str(page, row, const.column_sector),			# 公司大类
			"level": function.getcell_str(page, row, const.column_level),			# 公司层级
			"field": function.getcell_str(page, row, const.column_offer),			# 主要招聘学科
			"website1": function.getcell_str(page, row, const.column_website1),		# 官网
			"website2": function.getcell_str(page, row, const.column_website2),		# 校招官网
			"icon": function.getcell_str(page, row, const.column_icon),				# 图标1
			"images": function.getcell_str(page, row, const.column_images),			# 图标2
			"enttype": function.getcell_str(page, row, const.column_enttype),		# 央企国企
			"financial": function.getcell_str(page, row, const.column_financial),	# 金融机构
			"article1": function.getcell_str(page, row, const.column_article1),		# 深度解读
			"article2": function.getcell_str(page, row, const.column_article2),		# 招聘咨询
		}
		if enterprise['name'] == "深圳市创新投资集团有限公司":
			pass
		tag1 = function.getcell_str(page, row, const.column_tag1)					# 标签1
		tag2 = function.getcell_str(page, row, const.column_tag2)					# 标签2
		tag3 = function.getcell_str(page, row, const.column_tag3)					# 标签3
		tag4 = function.getcell_str(page, row, const.column_tag4)					# 标签4
		tag5 = function.getcell_str(page, row, const.column_tag5)					# 标签5
		tag = []
		if tag1 != "": tag.append(tag1)
		if tag2 != "": tag.append(tag2)
		if tag3 != "": tag.append(tag3)
		if tag4 != "": tag.append(tag4)
		if tag5 != "": tag.append(tag5)
		enterprise['tag'] = ','.join(tag)
		if enterprise['zone'] not in zone:
			print(f"地区不合法: {enterprise['zone']} 企业: {enterprise['name']}")
			continue
		if enterprise['sector'] not in sector:
			print(f"公司大类不合法: {enterprise['sector']} 企业: {enterprise['name']}")
			continue
		if enterprise['level'] not in level:
			print(f"公司层级不合法: {enterprise['level']} 企业: {enterprise['name']}")
			continue
		fields = enterprise['field'].split(',')
		valid = True
		valid_fields = []
		for f in fields:
			find = False
			if f.strip() == "": continue
			for field_item in field:
				if f.strip() in field_item['mapname']:
					find = True
					break
			if not find and f != "":
				print(f"主要招聘学科不合法: {f} 企业: {enterprise['name']}")
				valid = False
			else:
				valid_fields.append(f)
		if not valid: continue

		try:
			r = requests.post(function.url(const.url_insert_enterprise), json = enterprise, headers = const.request_headers, timeout = 15)
			r.raise_for_status()
			rj = json.loads(r.text)
			if 'code' not in rj or rj['code'] != 0:
				print(f"已失败: {enterprise['name']} response code {r.status_code} content: {r.text}")
		except requests.RequestException as err:
			print(f"上传失败 {enterprise['name']}: {err}")

def parse_case(page):
	for row in range(2, page.max_row + 1):
		case = {
			"name": function.getcell_str(page, row, 1),				# 岗位名称
			"enterprise": function.getcell_str(page, row, 2),		# 公司名称
			"field": function.getcell_str(page, row, 3),			# 专业名称
			"tags": function.getcell_str(page, row, 4),				# 标签
			"student": function.getcell_str(page, row, 5),			# 学生姓名
			"school1": function.getcell_str(page, row, 6),			# 本科院校
			"school1_tag": function.getcell_str(page, row, 7),		# 本科院校标签
			"field1": function.getcell_str(page, row, 8),			# 本科专业
			"school2": function.getcell_str(page, row, 9),			# 研究生院校
			"school2_tag": function.getcell_str(page, row, 10),		# 研究生院校标签
			"field2": function.getcell_str(page, row, 11),			# 研究生专业
			"year": function.getcell_int(page, row, 12),			# 毕业年份
			"detail": function.getcell_str(page, row, 13),			# 主要经历
		}
		try:
			r = requests.post(function.url(const.url_insert_case), json = case, headers = const.request_headers, timeout = 15)
			r.raise_for_status()
			rj = json.loads(r.text)
			if 'code' not in rj or rj['code'] != 0:
				print(f"已失败: {case['name']} response code {r.status_code} content: {r.text}")
		except requests.RequestException as err:
			print(f"上传失败 {case['name']}: {err}")

wb = openpyxl.load_workbook('企业列表.xlsx')
parse_global(wb['全局设置'])
parse_field(wb['学科列表'])
upload_config()
parse_enterprise(wb['第一批企业'])
parse_case(wb['成功案例'])

wb.close()

#truncate table qzj_zone;
#truncate table qzj_sector;
#truncate table qzj_level;
#truncate table qzj_field;
#truncate table qzj_enterprise;
#truncate table qzj_enterprise_field;
#truncate table qzj_case;
