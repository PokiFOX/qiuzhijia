from openpyxl.utils import column_index_from_string, get_column_letter
import unicodedata

from tapah import reserved

def url(path):
	return f'http://{reserved.backend_host}:{reserved.backend_port}/{path}'

def getcell(sheet, row, col):
	if isinstance(col, str):
		col = column_index_from_string(col)
	for merge in sheet.merged_cells:
		if row >= merge.min_row and row <= merge.max_row and col >= merge.min_col and col <= merge.max_col:
			return sheet.cell(merge.min_row, merge.min_col)
	return sheet.cell(row, col)

def getcell_int(sheet, row, col):
	cell = getcell(sheet, row, col)
	if cell.value == None:
		return 0
	elif type(cell.value) == str:
		v = str(cell.value).strip()
		if v == '':
			return 0
		if v[:2].lower() == '0x':
			return str2hex(v)
		return int(v)
	else:
		return int(cell.value)

def getcell_str(sheet, row, col):
	cell = getcell(sheet, row, col)
	if cell.value == None:
		return ''
	elif type(cell.value) == str:
		return cell.value
	else:
		return str(cell.value)

def getcell_bool(sheet, row, col):
	cell = getcell(sheet, row, col)
	if cell.value == '是':
		return True
	return False

def getcell_color(sheet, row, col):
	cell = getcell(sheet, row, col)
	if cell.font.color is None or cell.font.color.rgb is None:
		return ''
	return cell.font.color.rgb

def getcell_array(sheet, row, col):
	cell = getcell(sheet, row, col)
	if cell.value == None:
		return []
	elif type(cell.value) == int:
		return [cell.value]
	elif type(cell.value) != str:
		return []
	list = []
	for v in str(cell.value).split(','):
		if v.strip() != '':
			list.append(int(v.strip()))
		else:
			list.append(None)
	return list

def str2hex(str):
	str = str.upper()
	if len(str) > 2 and str[0:2] == '0X':
		str = str[2:]
	else:
		return int(str)
	result = 0
	for i in range(len(str)):
		c = str[len(str) - 1 - i]
		if c == '0':
			pass
		elif c == '1':
			result += 1 << (4 * i)
		elif c == '2':
			result += 2 << (4 * i)
		elif c == '3':
			result += 3 << (4 * i)
		elif c == '4':
			result += 4 << (4 * i)
		elif c == '5':
			result += 5 << (4 * i)
		elif c == '6':
			result += 6 << (4 * i)
		elif c == '7':
			result += 7 << (4 * i)
		elif c == '8':
			result += 8 << (4 * i)
		elif c == '9':
			result += 9 << (4 * i)
		elif c == 'A':
			result += 10 << (4 * i)
		elif c == 'B':
			result += 11 << (4 * i)
		elif c == 'C':
			result += 12 << (4 * i)
		elif c == 'D':
			result += 13 << (4 * i)
		elif c == 'E':
			result += 14 << (4 * i)
		elif c == 'F':
			result += 15 << (4 * i)
	return result

def format_page(page):
	def excel_length(text: str):
		total = 0
		for ch in text:
			if unicodedata.east_asian_width(ch) in ("W", "F"):
				total += 2
			else:
				total += 1
		return total

	width = []
	for row in page.iter_rows():
		for i, cell in enumerate(row):
			if len(width) <= i:
				width.append(0)
			if cell.value is not None:
				w = excel_length(str(cell.value))
				width[i] = max(width[i], w)

	for i, w in enumerate(width):
		page.column_dimensions[get_column_letter(i + 1)].width = w + 2
